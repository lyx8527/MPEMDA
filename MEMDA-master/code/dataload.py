import random

import numpy as np
import math
import torch
from openpyxl.reader.excel import load_workbook

from config_init import get_config

config = get_config()
dataset = config.dataset_topofallfeature
device = config.device_topofallfeature
path = fr"../dataset/{dataset}"
file_2 = load_workbook(fr"../dataset/{dataset}/microbe-disease({dataset}).xlsx")
sheet = file_2["Sheet1"]


def Amatrix2():
    pp = sheet.max_row
    qq = sheet.max_column
    view = np.empty((pp, qq))
    for i in range(1, pp + 1):
        for j in range(1, qq + 1):
            view[i - 1][j - 1] = sheet.cell(i, j).value
    n = int(view.max(axis=0)[0])
    m = int(view.max(axis=0)[1])
    A = np.zeros((n, m))
    for i in range(0, pp):
        A[int(view[i, 0]) - 1, int(view[i, 1]) - 1] = 1
    return A, pp, m, n


def A2matrix2_Drugvirus(A, partrandlist):
    A2 = A.copy()
    rand = np.zeros((len(partrandlist) + int(len(partrandlist)), 2))
    cnt = 0
    random.seed(1234)
    while cnt < int(len(partrandlist)):
        a = random.randint(0, A.shape[0] - 1)
        b = random.randint(0, A.shape[1] - 1)
        if A[a, b] != 1 and A2[a, b] != 1:
            rand[cnt, 0] = a
            rand[cnt, 1] = b
            cnt += 1
    for ii in partrandlist:
        A2[int(sheet.cell(ii + 1, 1).value - 1), int(sheet.cell(ii + 1, 2).value - 1)] = 0
        rand[cnt][0] = int(sheet.cell(ii + 1, 1).value - 1)
        rand[cnt][1] = int(sheet.cell(ii + 1, 2).value - 1)
        cnt = cnt + 1
    return A2, rand


# microbe similarity
def gausssim_microbe(A, m):
    sum = 0
    for i in range(0, m):
        sum += np.linalg.norm(A[:, i], ord=None, axis=0) ** 2
    gama = 1 / (sum / m)

    MS = np.empty((m, m))
    for i in range(0, m):
        for j in range(0, m):
            MS[i, j] = math.exp(-gama * (np.linalg.norm((A[:, i] - A[:, j]), ord=None) ** 2))
    return MS


def gausssim_microbe2(A, m):
    MS = np.empty((m, m))
    for i in range(0, m):
        for j in range(0, m):
            MS[i, j] = math.exp(-0.1 * (np.linalg.norm((A[:, i] - A[:, j]), ord=None) ** 2))
    return MS


# drug similarity
def gausssim_drug(A, n):
    sum = 0
    for i in range(0, n):
        sum += np.linalg.norm(A[i, :], ord=None) ** 2
    gama = 1 / (sum / n)

    DS = np.empty((n, n))
    for i in range(0, n):
        for j in range(0, n):
            DS[i, j] = math.exp(-gama * (np.linalg.norm((A[i, :] - A[j, :]), ord=None) ** 2))
    return DS


def cosinesim_microbe(A, m):
    MS = np.empty((m, m))
    for i in range(0, m):
        for j in range(0, m):
            MS[i, j] = np.dot(A[:, i], A[:, j]) / (
                    np.linalg.norm(A[:, i], ord=None) * np.linalg.norm(A[:, j], ord=None))
    MS[MS > 1] = 1
    return MS


def cosinesim_drug(A, n):
    DS = np.empty((n, n))
    for i in range(0, n):
        for j in range(0, n):
            DS[i, j] = np.dot(A[i, :], A[j, :]) / (
                    np.linalg.norm(A[i, :], ord=None) * np.linalg.norm(A[j, :], ord=None))
    DS[DS > 1] = 1
    return DS


def atcsim_drug():
    DD = np.loadtxt(path + '/Drug_atc.txt', dtype=np.float32)
    return DD


def sequencesim_microbe():
    MM = np.loadtxt(path + '/Microbe_sequence.txt', dtype=np.float32)
    return MM
